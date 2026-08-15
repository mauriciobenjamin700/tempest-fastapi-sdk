"""Row-by-row worksheet writing with a cursor, column specs and a theme.

``openpyxl`` addresses cells absolutely: every write is a ``(row, column)``
pair the caller has to track, and every styled cell repeats the same four
assignments (``font``, ``fill``, ``alignment``, ``border``). A document
built that way accumulates row arithmetic — insert a line near the top and
every constant below it shifts — and the styling drifts cell by cell.

:class:`SheetWriter` keeps the row cursor, applies per-column number
formats and alignment from a single :class:`Column` list, and paints the
theme from a :class:`~tempest_fastapi_sdk.spreadsheet.styles.SheetStyle`.
Callers describe the table once and then append rows.

    from decimal import Decimal
    from tempest_fastapi_sdk.spreadsheet import (
        BR_CURRENCY_FORMAT, Column, SheetWriter, new_workbook,
        workbook_to_bytes,
    )

    workbook = new_workbook("Orçamento")
    writer = SheetWriter(
        workbook["Orçamento"],
        columns=[
            Column("Item", width=48, wrap=True),
            Column("Qtd.", width=12, horizontal="center"),
            Column("Valor", width=18, number_format=BR_CURRENCY_FORMAT),
        ],
    )
    writer.title_block(["PREFEITURA MUNICIPAL", "Pregão 1/2026"])
    writer.header_row()
    writer.write_row(["Serviço de instalação", 2, Decimal("2930.00")])
    writer.total_row(["Total", None, Decimal("5860.00")])
    writer.apply_widths()
    data = workbook_to_bytes(workbook)

Needs the ``[spreadsheet]`` extra (``openpyxl``). The engine is imported at
first use, so importing this module — and defining a project's columns and
theme — works without it.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING, Any

from tempest_fastapi_sdk.spreadsheet.styles import DEFAULT_SHEET_STYLE, SheetStyle

if TYPE_CHECKING:
    from collections.abc import Sequence

    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

CellValue = str | int | float | Decimal | date | datetime | None
"""What a cell may receive.

Narrower than what ``openpyxl`` accepts (it also takes ``bytes``, rich text
and formula objects) and deliberately so — widening it to ``object`` is what
lets an unconverted domain object reach a cell and land in the file as its
``repr``.

A ``str`` starting with ``=`` is written as a live formula by ``openpyxl``,
which is how a closing check can react to an edited value instead of
freezing the number that was true at generation time.
"""

_MISSING_DEPENDENCY = (
    "openpyxl is required to write spreadsheets. Install the extra: "
    'pip install "tempest-fastapi-sdk[spreadsheet]"'
)


@dataclass(frozen=True, slots=True)
class Column:
    """Specification of one column: its header, width and cell treatment.

    Declared once and reused for every row, so a column's number format
    cannot drift between the first row and the thousandth.

    Attributes:
        title (str): Header text.
        width (int | None): Column width in characters. ``None`` leaves the
            workbook default, which truncates long text on screen.
        number_format (str | None): Excel number format applied to every
            body cell of this column — see
            :mod:`tempest_fastapi_sdk.spreadsheet.formats`. ``None`` leaves
            the value as Excel guesses it.
        horizontal (str | None): Horizontal alignment (``"left"``,
            ``"center"``, ``"right"``). ``None`` uses Excel's default, which
            already right-aligns numbers and left-aligns text.
        wrap (bool): Whether body cells wrap their text. Turn it on for the
            description column and leave it off elsewhere: a wrapped short
            cell is what makes a row unnecessarily tall.
    """

    title: str
    width: int | None = None
    number_format: str | None = None
    horizontal: str | None = None
    wrap: bool = False


class SheetWriter:
    """Append rows to a worksheet, styled from column specs and a theme.

    Stateful in exactly one respect — :attr:`row`, the next free row — so
    two writers may share a workbook but never a sheet.

    Attributes:
        sheet (Worksheet): The worksheet being written.
        columns (tuple[Column, ...]): Column specifications, left to right.
        style (SheetStyle): The theme applied to headers, groups and totals.
        row (int): The next free row, 1-based. Assign to it to jump.
    """

    def __init__(
        self,
        sheet: Worksheet,
        columns: Sequence[Column] = (),
        *,
        style: SheetStyle = DEFAULT_SHEET_STYLE,
        start_row: int = 1,
    ) -> None:
        """Bind a writer to a worksheet.

        Args:
            sheet (Worksheet): The sheet to write into.
            columns (Sequence[Column]): Column specifications, left to
                right. May be empty when the sheet is a free-form block of
                label/value rows rather than a table.
            style (SheetStyle): Colours and sizes for emphasised rows.
            start_row (int): First row to write, 1-based.
        """
        self.sheet = sheet
        self.columns: tuple[Column, ...] = tuple(columns)
        self.style = style
        self.row = start_row

    @property
    def column_count(self) -> int:
        """How many columns the sheet spans.

        Returns:
            int: The number of declared columns, or ``1`` when none were
            declared — a title block still has to merge across something.
        """
        return len(self.columns) or 1

    def title_block(
        self,
        lines: Sequence[str],
        *,
        span: int | None = None,
    ) -> int:
        """Write centred, merged title lines across the table width.

        The identification block a formal document opens with: issuing body,
        process number, subject. Each line is merged across ``span``
        columns, so the block stays centred when a column is added.

        Args:
            lines (Sequence[str]): One string per line, top to bottom. Empty
                strings are written as empty merged rows, which is how a
                caller leaves deliberate breathing room inside the block.
            span (int | None): How many columns to merge across. Defaults to
                :attr:`column_count`.

        Returns:
            int: The next free row.
        """
        from openpyxl.utils import get_column_letter

        last_column = get_column_letter(span or self.column_count)
        font = self._font(bold=True, size=self.style.title_size)
        alignment = self._alignment(horizontal="center", wrap=True)
        for text in lines:
            self.sheet.merge_cells(f"A{self.row}:{last_column}{self.row}")
            cell = self.sheet.cell(row=self.row, column=1, value=text)
            cell.font = font
            cell.alignment = alignment
            self.row += 1
        return self.row

    def header_row(self, titles: Sequence[str] | None = None) -> int:
        """Write the styled column-header row.

        Args:
            titles (Sequence[str] | None): Header texts. ``None`` uses the
                titles from :attr:`columns`, which is the point of declaring
                them; pass an explicit list only for a second header band.

        Returns:
            int: The next free row.
        """
        from openpyxl.styles import PatternFill

        texts = list(titles) if titles is not None else [c.title for c in self.columns]
        font = self._font(
            bold=True,
            size=self.style.header_size,
            color=self.style.header_foreground,
        )
        fill = PatternFill("solid", fgColor=self.style.header_background)
        alignment = self._alignment(horizontal="center", wrap=True)
        border = self._border()
        for index, text in enumerate(texts, start=1):
            cell = self.sheet.cell(row=self.row, column=index, value=text)
            cell.font = font
            cell.fill = fill
            cell.alignment = alignment
            cell.border = border
        self.row += 1
        return self.row

    def write_row(
        self,
        values: Sequence[CellValue],
        *,
        formats: Sequence[str | None] | None = None,
        bold: bool = False,
        background: str | None = None,
        border: bool = True,
    ) -> int:
        """Write one body row, applying each column's format and alignment.

        Args:
            values (Sequence[CellValue]): Cell values, left to right. A
                value of ``None`` leaves the cell empty but still styled, so
                a total row can skip the middle columns and keep its fill.
                Values may be shorter or longer than :attr:`columns`;
                columns beyond the declared ones get no format.
            formats (Sequence[str | None] | None): Per-cell number formats
                overriding the column's, for the row where one cell holds a
                percentage in an otherwise monetary column.
            bold (bool): Whether to bold the row.
            background (str | None): Hex fill for the whole row. ``None``
                leaves it unfilled.
            border (bool): Whether to draw the thin grid around the cells.

        Returns:
            int: The next free row.
        """
        from openpyxl.styles import PatternFill

        fill = PatternFill("solid", fgColor=background) if background else None
        edge = self._border() if border else None
        for index, value in enumerate(values, start=1):
            column = self.columns[index - 1] if index <= len(self.columns) else None
            cell = self.sheet.cell(row=self.row, column=index, value=value)
            cell.font = self._font(bold=bold, size=self.style.header_size)
            if fill is not None:
                cell.fill = fill
            if edge is not None:
                cell.border = edge
            number_format = self._format_for(index, column, formats)
            if number_format is not None:
                cell.number_format = number_format
            if column is not None and (column.horizontal or column.wrap):
                cell.alignment = self._alignment(
                    horizontal=column.horizontal,
                    wrap=column.wrap,
                )
        self.row += 1
        return self.row

    def group_row(self, values: Sequence[CellValue]) -> int:
        """Write a section heading inside the table.

        Args:
            values (Sequence[CellValue]): Cell values, usually just the
                group label in the first column.

        Returns:
            int: The next free row.
        """
        return self.write_row(
            values,
            bold=True,
            background=self.style.group_background,
        )

    def total_row(self, values: Sequence[CellValue]) -> int:
        """Write an emphasised total row.

        Args:
            values (Sequence[CellValue]): Cell values, usually a label and
                one or more amounts.

        Returns:
            int: The next free row.
        """
        return self.write_row(
            values,
            bold=True,
            background=self.style.total_background,
        )

    def blank_rows(self, count: int = 1) -> int:
        """Advance the cursor past empty rows, writing nothing.

        Args:
            count (int): How many rows to skip.

        Returns:
            int: The next free row.
        """
        self.row += count
        return self.row

    def apply_widths(self) -> None:
        """Set the width of every column that declared one.

        Call it once, after the rows are written or before — widths are a
        property of the column, not of the content.
        """
        from openpyxl.utils import get_column_letter

        for index, column in enumerate(self.columns, start=1):
            if column.width is not None:
                letter = get_column_letter(index)
                self.sheet.column_dimensions[letter].width = column.width

    def freeze_below(self, row: int | None = None) -> None:
        """Freeze every row above ``row`` so headers survive scrolling.

        Args:
            row (int | None): First scrolling row. ``None`` freezes
                everything written so far, which right after
                :meth:`header_row` is exactly the header band.
        """
        self.sheet.freeze_panes = f"A{row if row is not None else self.row}"

    def _format_for(
        self,
        index: int,
        column: Column | None,
        formats: Sequence[str | None] | None,
    ) -> str | None:
        """Resolve the number format for one cell.

        Args:
            index (int): 1-based column index.
            column (Column | None): The column spec, when declared.
            formats (Sequence[str | None] | None): Per-row overrides.

        Returns:
            str | None: The format to apply, or ``None`` to leave the cell
            as Excel guesses it. A per-row override wins over the column's,
            including when it is explicitly ``None``.
        """
        if formats is not None and index <= len(formats):
            return formats[index - 1]
        return column.number_format if column is not None else None

    def _font(
        self,
        *,
        bold: bool = False,
        size: int | None = None,
        color: str | None = None,
    ) -> Any:
        """Build a ``Font`` from the theme.

        Args:
            bold (bool): Whether the text is bold.
            size (int | None): Point size; ``None`` uses the workbook
                default.
            color (str | None): Hex text colour; ``None`` uses the default.

        Returns:
            Any: An ``openpyxl.styles.Font``. Typed as ``Any`` because the
            engine is imported lazily and must not appear in this module's
            import-time namespace.
        """
        from openpyxl.styles import Font

        return Font(bold=bold, size=size, color=color, name=self.style.font_name)

    def _alignment(
        self,
        *,
        horizontal: str | None = None,
        wrap: bool = False,
    ) -> Any:
        """Build an ``Alignment`` with the vertical anchoring a table wants.

        Vertical is always ``top``: with wrapped text, centring makes rows of
        different heights read as if they were misaligned.

        Args:
            horizontal (str | None): Horizontal alignment.
            wrap (bool): Whether to wrap text.

        Returns:
            Any: An ``openpyxl.styles.Alignment``.
        """
        from openpyxl.styles import Alignment

        return Alignment(horizontal=horizontal, vertical="top", wrap_text=wrap)

    def _border(self) -> Any:
        """Build the thin grid border from the theme.

        Returns:
            Any: An ``openpyxl.styles.Border`` with all four thin sides.
        """
        from openpyxl.styles import Border, Side

        side = Side(style="thin", color=self.style.border_color)
        return Border(left=side, right=side, top=side, bottom=side)


def new_workbook(*sheet_titles: str) -> Workbook:
    """Create a workbook holding exactly the named sheets.

    ``openpyxl`` always creates a workbook with one sheet called ``Sheet``.
    Forgetting to remove it ships a document with a stray empty tab, which
    is the kind of detail that makes a generated file look generated.

    Args:
        *sheet_titles (str): Sheet names, in tab order. Passing none returns
            a workbook with the default sheet untouched.

    Returns:
        Workbook: The workbook, with the sheets created and the default one
        removed when titles were given.

    Raises:
        ImportError: When the ``[spreadsheet]`` extra is not installed.
    """
    try:
        from openpyxl import Workbook as WorkbookClass
    except ImportError as exc:  # pragma: no cover - exercised without the extra
        raise ImportError(_MISSING_DEPENDENCY) from exc

    workbook = WorkbookClass()
    if not sheet_titles:
        return workbook
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)
    for title in sheet_titles:
        workbook.create_sheet(title)
    return workbook


def workbook_to_bytes(workbook: Workbook) -> bytes:
    """Serialize a workbook to ``.xlsx`` bytes in memory.

    Keeps the file off disk, which is what lets a handler stream it as a
    download or hand it to object storage without a temporary path to clean
    up (and without a race between two requests writing the same one).

    Args:
        workbook (Workbook): The workbook to serialize.

    Returns:
        bytes: The ``.xlsx`` file contents.
    """
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


__all__: list[str] = [
    "CellValue",
    "Column",
    "SheetWriter",
    "new_workbook",
    "workbook_to_bytes",
]
